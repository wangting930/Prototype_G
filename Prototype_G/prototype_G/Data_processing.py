import numpy as np
import openpyxl
import random
from collections import namedtuple


Data = namedtuple('Data', ['x', 'y', 'adjacency_dict',
                           'train_mask', 'val_mask', 'test_mask'])


def process_data():

        #Data processing

        pre_data = openpyxl.load_workbook('Location of the data')
        table = pre_data.active
        nrows = table.max_row  # Get the number of valid rows in the sheet
        tezheng_dim=1024   #feature input dimension
        x = np.zeros((nrows, tezheng_dim))
        y=np.arange(nrows)

        adjacency_dict = np.full((nrows, nrows+1), 0, dtype=int)
        print("Process data ...")
        print(nrows)
        for i in range(nrows):
            if table.cell(i + 1, 3).value==0:
                y[i]=0
            else:
                if table.cell(i + 1, 3).value==1:
                    y[i]=1
                else:
                    if table.cell(i + 1, 3).value==2:
                        y[i] = 2
                    else:
                        if table.cell(i + 1, 3).value == 3:
                            y[i] = 3
                        else:
                            if table.cell(i + 1, 3).value == 4:
                                y[i] = 4
            adjacency_dict[i][0] = table.cell(i + 1, 1).value
            for j in range(nrows):
                if table.cell(i + 1, j + tezheng_dim+4).value == 1:
                    adjacency_dict[i][j+1] = 1
            for m in range(tezheng_dim):
                x[i][m] = table.cell(i + 1, m + 4).value

        #Train set and test set

        test_index = random.sample(range(0,600), 100)
        random.shuffle(test_index)
        train_index=[]
        val_index = np.arange(320,320)
        for j in range(nrows):
            if j not in test_index:
                train_index.append(j)
        random.shuffle(train_index)

        num_nodes = x.shape[0]

        train_mask = np.zeros(num_nodes, dtype=np.bool)
        val_mask = np.zeros(num_nodes, dtype=np.bool)
        test_mask = np.zeros(num_nodes, dtype=np.bool)
        train_mask[train_index] = True
        val_mask[val_index] = True
        test_mask[test_index] = True
        print("Node's feature shape: ", x.shape)
        print("Node's label shape: ", y.shape)
        print("Adjacency's shape: ", len(adjacency_dict))
        print("Number of training nodes: ", train_mask.sum())
        print("Number of validation nodes: ", val_mask.sum())
        print("Number of test nodes: ", test_mask.sum())

        return Data(x=x, y=y, adjacency_dict=adjacency_dict,
                    train_mask=train_mask, val_mask=val_mask,test_mask=test_mask)
