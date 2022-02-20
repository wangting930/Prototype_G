import openpyxl
import numpy as np

data = openpyxl.load_workbook('.xlsx')
table = data.active#通过索引顺序获取
nrows = table.max_row#获取该sheet中的有效行数
print(nrows)
array = np.zeros((nrows,4096))
a=np.zeros((nrows,nrows))
bian=np.full((nrows,nrows), 0, dtype=int)
d=np.full((nrows), 100, dtype=int)
for i in range(nrows):
    for j in range(4096):
        array[i][j] =table.cell(i+1, j+4).value

def cos_sim(a, b):
    a_norm = np.linalg.norm(a)
    b_norm = np.linalg.norm(b)
    cos = np.dot(a,b)/(a_norm * b_norm)
    return cos

for i in range(nrows):
    for j in range(nrows):
        if i==j:
            a[i][j]=0
        else:
            a[i][j]=cos_sim(array[i],array[j])
for i in range(nrows):
    for j in range(nrows):
        if d[i]>0:
            b = np.argmax(a[i])
            if d[b]>0:
                bian[i][b] = 1
                bian[b][i] = 1
                d[b]=d[b]-1
                a[i][b] = 0
                d[i]=d[i]-1
            else:
                a[i][b]=0
                continue


print(d)
for i in range(nrows):
   for j in range(nrows):
        table.cell(i+1,j+4096+4).value = bian[i][j]
      #data.save("D:\stop的各种文档啊\其余的论文啊\wt.csv")#将边进行保存
data.save(".xlsx")#将边进行保存
print("已完成")