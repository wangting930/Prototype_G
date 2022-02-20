import openpyxl
import numpy as np

data = openpyxl.load_workbook('D:\\.xlsx')
table = data.active#通过索引顺序获取
nrows = table.max_row#获取该sheet中的有效行数
print(nrows)
array = np.zeros((nrows,1024))
a=np.zeros((nrows,nrows))
bian=np.full((nrows,nrows), 0, dtype=int)
d=np.full((nrows), 100, dtype=int)
for i in range(nrows):
    for j in range(1024):
        array[i][j] =table.cell(i+1, j+4).value


for i in range(nrows):
    for j in range(nrows):
        if i==j:
            a[i][j]=1000
        else:
            a[i][j]=np.sqrt(np.sum(np.square(array[i]-array[j])))


for i in range(nrows):
    for j in range(nrows):
        if i==j:
            bian[i][j]==0
        if d[i]>0:
            b = np.argmin(a[i])
            if d[b]>0:
                bian[i][b] = 1
                bian[b][i] = 1
                d[b]=d[b]-1
                a[i][b] = 1000
                d[i]=d[i]-1
            else:
                a[i][b]=1000
                continue


print(d)
for i in range(nrows):
   for j in range(nrows):
        table.cell(i+1,j+1028).value = bian[i][j]
      #将边进行保存
data.save("D:\.xlsx")#将边进行保存
print("已完成")